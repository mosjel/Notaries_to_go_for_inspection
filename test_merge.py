import pandas as pd
from openpyxl import load_workbook
import numpy as np
import jdatetime
import openpyxl

def persian_to_georgian(persin_date):
    try:
        if persin_date!="" or None:
            date_list=persin_date.split("/")
            date_list=[int(i) for i in date_list]
            gregorian_date=jdatetime.date(date_list[0],date_list[1],date_list[2]).togregorian()
            return(gregorian_date)
    except Exception as e:
        return "Error"
def forward_fill_merged(df, merged_ranges):
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(':')
       

        
        # Convert cell addresses to indices
        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        end_row, end_col = openpyxl.utils.cell.coordinate_to_tuple(end_cell)

        # Adjust for zero-based indexing in pandas
        start_row -= 2
        end_row -= 1
        start_col -=1  
        # Select the range in the DataFrame
        sub_df = df.iloc[start_row:end_row, start_col:end_col]
       
        # Apply forward fill to the range
        sub_df = sub_df.ffill(axis=0)
        
        # Put the filled values back into the original DataFrame
        df.iloc[start_row:end_row, start_col:end_col] = sub_df

file_path = r"C:\Users\VAIO\Desktop\test\book1.xlsx"
workbook = openpyxl.load_workbook(file_path)
 
sheet = workbook['Sheet1']  # Select the sheet named "Sheet1"

# List to store the addresses of merged cells
merged_cells_ranges = []

# Iterate through the merged cells ranges
for merged_range in sheet.merged_cells.ranges:
    merged_cells_ranges.append(merged_range.coord)
column_dype={4:str}
df = pd.read_excel(file_path, engine='openpyxl',sheet_name="Sheet1",dtype=column_dype)
df[df.columns[4]]=df[df.columns[4]].fillna("")
print(df)
print (df.index.tolist())
input("-----------------------------------222222222222222222222222222222222222222222222")


# Forward fill to propagate the merged values down the rows
forward_fill_merged(df, merged_cells_ranges)
# df.ffill(inplace=True)
# print(df.shape)
print(df,"----------------------------------*************************")
input()

# df_merged = df.groupby(df.columns[[0,1,2,3]].tolist(),dropna=False)[df.columns[4]].apply(lambda x: ('*'.join(x))).reset_index()
grouped= df.groupby(df.columns[[0,1,2,3]].tolist(),dropna=False).apply(lambda x:x.index.min()+2).reset_index()
print(grouped)
print("JJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJJ")
df_merged = df.groupby(df.columns[[0,1,2,3]].tolist(),dropna=False)[df.columns[4]].apply(lambda x:  ('*'.join(x))).reset_index()
df_merged["Excel_Index"]=grouped[grouped.columns[4]]
# df_merged = df.groupby(['ردیف', 'نام', 'تاریخ'])['موارد'].apply(lambda x: ' '.join(x)).reset_index()

# Display the resulting DataFrame
print(df_merged)
print("3333333333333333333333333333333333333333")





# df_rule_cont=df_merged[df_merged[df.columns[5]] != "تخلف ندارد"]

# df_rule_cont[df_rule_cont.columns[2]]=df_rule_cont[df_rule_cont.columns[2]].fillna("")


# df_rule_cont["georgian"]=df_rule_cont[df_rule_cont.columns[2]].apply (lambda x:persian_to_georgian(x))
# print(df_rule_cont)



# output_file_path=r"C:\Users\VAIO\Desktop\test\Book2.xlsx"
# df_rule_cont.to_excel(output_file_path,index=False)

# # with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
# #     # Write the new data to "Sheet2"
# #     new_data.to_excel(writer, sheet_name='Sheet2', index=False)
# print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
# workbook = openpyxl.load_workbook(output_file_path)
# sheet = workbook.active

# # Set the sheet to RTL mode
# sheet.sheet_view.rightToLeft = True

# # Save the workbook with the RTL modification
# workbook.save(output_file_path)
































# df_rule_cont[df_rule_cont.columns[7]].fillna(pd.NaT, inplace=True)

# df_error=df_rule_cont[df_rule_cont[df_rule_cont.columns[7]]=="Error"]
# print(df_error)
# input("""""")








# df_rule_cont[df_rule_cont.columns[7]] = pd.to_datetime(df_rule_cont[df_rule_cont.columns[7]], errors='coerce').dt.date
# print(df_rule_cont)
# input("**************************")





# max_dates = df_rule_cont.groupby(df_rule_cont.columns[4]) [df_rule_cont.columns[7]].max()















# df_rule_cont_max_date=df_rule_cont[df_rule_cont.groupby(df.columns[4],dropna=False)[df_rule_cont.columns[7]].transform('max') == df_rule_cont[df_rule_cont.columns[7]]]
# print(df_rule_cont_max_date)
# df_rule_cont_max_date=df_rule_cont_max_date.sort_values(by=df_rule_cont_max_date.columns[7],ascending=True)
# print(df_rule_cont_max_date)
# output_file_path=r"C:\Users\VAIO\Desktop\test\Book2.xlsx"
# df_rule_cont_max_date.to_excel(output_file_path,index=False)

# # with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
# #     # Write the new data to "Sheet2"
# #     new_data.to_excel(writer, sheet_name='Sheet2', index=False)
# print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
# workbook = openpyxl.load_workbook(output_file_path)
# sheet = workbook.active

# # Set the sheet to RTL mode
# sheet.sheet_view.rightToLeft = True

# # Save the workbook with the RTL modification
# workbook.save(output_file_path)