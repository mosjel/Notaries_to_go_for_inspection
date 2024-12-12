import pandas as pd
from openpyxl import load_workbook
import numpy as np
import jdatetime
import openpyxl

def persian_to_georgian(persin_date):
    try:
        if persin_date!="" or None:
            date_list=persin_date.split("/")
            date_list=[int(i) for i in date_list]
            gregorian_date=jdatetime.date(date_list[0],date_list[1],date_list[2]).togregorian()
            return(gregorian_date)
    except Exception as e:
        return "Error"
def forward_fill_merged(df, merged_ranges):
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(':')
       

        
        # Convert cell addresses to indices
        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        end_row, end_col = openpyxl.utils.cell.coordinate_to_tuple(end_cell)

        # Adjust for zero-based indexing in pandas
        start_row -= 2
        end_row -= 1
        start_col -=1  
        # Select the range in the DataFrame
        sub_df = df.iloc[start_row:end_row, start_col:end_col]
       
        # Apply forward fill to the range
        sub_df = sub_df.ffill(axis=0)
        
        # Put the filled values back into the original DataFrame
        df.iloc[start_row:end_row, start_col:end_col] = sub_df

file_path = r"C:\Users\VAIO\Desktop\test\total.xlsx"
workbook = openpyxl.load_workbook(file_path)
 
sheet = workbook['Sheet1']  # Select the sheet named "Sheet1"

# List to store the addresses of merged cells
merged_cells_ranges = []

# Iterate through the merged cells ranges
for merged_range in sheet.merged_cells.ranges:
    merged_cells_ranges.append(merged_range.coord)
column_dtype = {6: str,4:str}
df = pd.read_excel(file_path, engine='openpyxl',usecols=list(range(0,7)),sheet_name="Sheet1",dtype=column_dtype)
df[df.columns[6]]=df[df.columns[6]].fillna("")
last_valid_index=df.iloc[:,6].last_valid_index()
print("Last Row is:",last_valid_index+2)
df=df.iloc[:last_valid_index+1,:]
print(df)


# Forward fill to propagate the merged values down the rows
forward_fill_merged(df, merged_cells_ranges)
# df.ffill(inplace=True)
# print(df.shape)

grouped= df.groupby(df.columns[[0,1,2,3]].tolist(),dropna=False).apply(lambda x:x.index.min()+2).reset_index()
# df_merged = df.groupby(df.columns[[0,1,2,3]].tolist(),dropna=False)[df.columns[4]].apply(lambda x: ('*'.join(x))).reset_index()
df_merged = df.groupby(df.columns[[0,1,2,3,4,5]].tolist(),dropna=False)[df.columns[6]].apply(lambda x: ('*'.join(x))).reset_index()
df_merged["Excel_Index"]=grouped[grouped.columns[4]]
# df_merged = df.groupby(['ردیف', 'نام', 'تاریخ'])['موارد'].apply(lambda x: ' '.join(x)).reset_index()

# Display the resulting DataFrame


visited_notaries=np.array(df_merged[df_merged.columns[4]])

df_notary=pd.read_excel(r"C:\Users\VAIO\Desktop\test\Notary_list.xlsx",dtype={0:str})
notary_list=np.array(df_notary["لیست دفترخانه های تهران"])
notary_list_not_visited_index=np.invert(np.isin(notary_list,visited_notaries))

notary_visited_misakes_index=np.invert(np.isin(visited_notaries,notary_list))


































df_rule_cont=df_merged[df_merged[df.columns[5]] !="عدم تخلف"]

print(df_rule_cont)




















df_rule_cont[df_rule_cont.columns[2]]=df_rule_cont[df_rule_cont.columns[2]].fillna("")


df_rule_cont["georgian"]=df_rule_cont[df_rule_cont.columns[2]].apply (lambda x:persian_to_georgian(x))
print(df_rule_cont)



df_error=df_rule_cont[df_rule_cont[df_rule_cont.columns[8]]=="Error"]
df_na=df_rule_cont[df_rule_cont[df_rule_cont.columns[8]].isnull()]

df_rule_cont[df_rule_cont.columns[8]]=pd.to_datetime(df_rule_cont[df_rule_cont.columns[8]],errors="coerce").dt.date

df_rule_cont.dropna(subset=[df_rule_cont.columns[8]], inplace=True)


df_dates=df_rule_cont[df_rule_cont.groupby(df_rule_cont.columns[4],dropna=False)[df_rule_cont.columns[8]].transform("max")==df_rule_cont[df_rule_cont.columns[8]]]
df_dates=df_dates.sort_values(df_dates.columns[8],ascending=True)




df_merged["شماره دفترخانه اشتباه"]=""
df_merged.loc[notary_visited_misakes_index,"شماره دفترخانه اشتباه"]="Mistake Risk"
pd.set_option('display.max_columns', None)
df_mistake_risk=df_merged[notary_visited_misakes_index]
notary_list_not_visited=notary_list[notary_list_not_visited_index]
empty_df= pd.DataFrame('', index=range(len(notary_list_not_visited)), columns=df_mistake_risk.columns)
empty_df[empty_df.columns[4]]=notary_list_not_visited
print((empty_df).shape)
print((df_mistake_risk).shape)

input()
df_mistake_and_non=pd.concat([df_mistake_risk,empty_df])
print(df_mistake_and_non)
input()

df_org=pd.concat([df_mistake_and_non,df_error,df_na,df_dates])





print(df_org.shape)

df_org.insert(loc=0,column="ردیف اصلی",value=list(range(1,df_org.shape[0]+1)))
df_org.rename(columns=({df_org.columns[1]:"ردیف فایل مرجع"}),inplace=True)
output_file_path=r"C:\Users\VAIO\Desktop\test\Book2.xlsx"
df_org.to_excel(output_file_path,index=False)

# with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#     # Write the new data to "Sheet2"
#     new_data.to_excel(writer, sheet_name='Sheet2', index=False)
print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
workbook = openpyxl.load_workbook(output_file_path)
sheet = workbook.active

# Set the sheet to RTL mode
from openpyxl.styles import Alignment

sheet.sheet_view.rightToLeft = True

max_column_index = sheet.max_column
from openpyxl.utils import get_column_letter
# Set the width of all active columns to 13
for column_index in range(1, max_column_index + 1):
    column_letter = get_column_letter(column_index)
    sheet.column_dimensions[column_letter].width = 16

sheet.column_dimensions["H"].width = 40
from openpyxl.styles import PatternFill

red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Iterate over all cells in the specified column
for col_index in range(9, 11):
    for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index, max_row=sheet.max_row):
        for cell in row:
            # Check if cell value is "Error" or "Mistake Risk"
            if cell.value and cell.value.strip() in ["Error", "Mistake Risk"]:
                cell.fill = red_fill

for row in sheet.iter_rows():
    for cell in row:
        # Set text direction to right-to-left by setting reading order
        cell.alignment = Alignment(wrap_text=True,readingOrder=2,vertical="center",horizontal="center")

# Save the changes to the workbook

# Save the workbook with the RTL modification
workbook.save(output_file_path)



