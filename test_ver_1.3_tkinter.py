import tkinter as tk
from tkinter import messagebox, font
import pandas as pd
from openpyxl import load_workbook
import openpyxl.utils.cell
import numpy as np
import jdatetime
import os

# Function to handle the Persian text introduction
def introduction(root, process_label):
    s = [
        "--------------------تندر افزار نرم--------------------",
        "(بازرسی جهت رسمی اسناد دفاتر نوبت تعیین)",
        "1403 بهار - 1.2 نسخه",
        "نسخه 1.1",
        "-------------------------",
        "جلوه مصطفی",
        "Email: mostafaajelveh@gmail.com"
    ]

    # Define the number of columns (characters width) for centering text
    columns = 50  # You can adjust this value as needed

    # Create a Text widget for multiline display
    text_widget = tk.Text(root, wrap=tk.WORD, bg='black', fg='white', font=('Arial', 14), relief='flat')
    text_widget.pack(expand=True, fill='both')

    for num, i in enumerate(s):
        if num < 5:
            text_widget.insert(tk.END, i.center(columns, '-') + '\n', 'green_bold')
        else:
            text_widget.insert(tk.END, i.center(columns, '-') + '\n', 'white_bold')

    text_widget.insert(tk.END, '\n' * 4)

    for i in range(7):
        text_widget.insert(tk.END, '*' * columns + '\n', 'red_bold')

    text_widget.insert(tk.END, '\n\n' + process_label.center(columns) + '\n', 'blue_bold')

    # Configure tags for styling
    text_widget.tag_config('green_bold', foreground='#00FF00', font=('Arial', 14, 'bold'))
    text_widget.tag_config('white_bold', foreground='#FFFFFF', font=('Arial', 14, 'bold'))
    text_widget.tag_config('red_bold', foreground='#FF0000', font=('Arial', 14, 'bold'))
    text_widget.tag_config('blue_bold', foreground='#0000FF', font=('Arial', 14, 'bold'))

    # Disable editing in the Text widget
    text_widget.config(state='disabled')

    root.after(5000, lambda: root.destroy())  # Close window after 5 seconds

# Convert Persian date to Gregorian date
def persian_to_georgian(persian_date):
    try:
        if persian_date:
            date_list = [int(i) for i in persian_date.split("/")]
            gregorian_date = jdatetime.date(date_list[0], date_list[1], date_list[2]).togregorian()
            return gregorian_date
    except Exception as e:
        return "Error"

# Forward fill merged cells in a DataFrame
def forward_fill_merged(df, merged_ranges):
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(':')

        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        end_row, end_col = openpyxl.utils.cell.coordinate_to_tuple(end_cell)

        start_row -= 2
        end_row -= 1
        start_col -= 1

        sub_df = df.iloc[start_row:end_row, start_col:end_col]
        sub_df = sub_df.ffill(axis=0)
        df.iloc[start_row:end_row, start_col:end_col] = sub_df

# Main application logic
try:
    dir_name = os.path.dirname(os.path.abspath(__file__))
    os.system('cls' if os.name == 'nt' else 'clear')

    error_label = ": شد مواجه خطا با فایل ایجاد"
    process_label = "......[notary_visited_total.xlsx] فایل اطلاعات پردازش حال در"
    success_label = ".شد ایجاد موفقیت با [notaries_togo.xlsx]  فایل"
    continue_label = ".بفشارید را c دکمه انصراف برای یا و دهید فشار را Enter ادامه برای .شوید مطمئن جاری مسیر در [total.xlsx] فایل بودن موجود از ادامه از قبل "
    success_label_1 = u'\u2713' + " " + "باشید موفق"

    # Initialize Tkinter root window
    root = tk.Tk()
    root.title("تندر افزار نرم")
    root.geometry("800x600")

    # Call introduction function to display messages
    introduction(root, process_label)
    root.mainloop()

    file_path = os.path.join(dir_name, "total.xlsx")
    workbook = load_workbook(file_path)
    sheet = workbook['Sheet1']

    merged_cells_ranges = [merged_range.coord for merged_range in sheet.merged_cells.ranges]
    column_dtype = {6: str, 4: str}
    df = pd.read_excel(file_path, engine='openpyxl', usecols=list(range(0, 7)), sheet_name="Sheet1", dtype=column_dtype)
    df[df.columns[6]] = df[df.columns[6]].fillna("")
    last_valid_index = df.iloc[:, 6].last_valid_index()
    print("Last Row is:", last_valid_index + 2)
    df = df.iloc[:last_valid_index + 1, :]

    forward_fill_merged(df, merged_cells_ranges)

    grouped = df.groupby(df.columns[[0, 1, 2, 3, 4, 5]].tolist(), dropna=False).apply(lambda x: x.index.min() + 2).reset_index()
    df_merged = df.groupby(df.columns[[0, 1, 2, 3, 4, 5]].tolist(), dropna=False)[df.columns[6]].apply(lambda x: ('*'.join(x))).reset_index()
    df_merged["Excel_Index"] = grouped[grouped.columns[6]]

    visited_notaries = np.array(df_merged[df_merged.columns[4]])

    df_notary = pd.read_excel(os.path.join(dir_name, "notary_list.xlsx"), dtype={0: str})
    notary_list = np.array(df_notary["لیست دفترخانه های تهران"])
    notary_list_not_visited_index = np.invert(np.isin(notary_list, visited_notaries))
    notary_mistakes_index = np.invert(np.isin(visited_notaries, notary_list))

    df_merged[df_merged.columns[2]] = df_merged[df_merged.columns[2]].fillna("")

    print(df_merged)
    df_merged["georgian"] = df_merged[df_merged.columns[2]].apply(lambda x: persian_to_georgian(x))
    print(df_merged)

    df_error = df_merged[df_merged[df_merged.columns[8]] == "Error"]
    df_na = df_merged[df_merged[df_merged.columns[8]].isnull()].copy()
    df_na[df_na.columns[2]] = "Mistake Risk"
    df_na_violated = df_na[df_na[df_na.columns[5]] != "عدم تخلف"]
    df_na_not_violated = df_na[df_na[df_na.columns[5]] == "عدم تخلف"]

    df_merged[df_merged.columns[8]] = pd.to_datetime(df_merged[df_merged.columns[8]], errors="coerce").dt.date
    df_mistake_risk = df_merged[notary_mistakes_index].copy()
    df_mistake_risk["شماره دفترخانه اشتباه"] = "Mistake"
    df_merged.dropna(subset=[df_merged.columns[8]], inplace=True)

    df_dates = df_merged[df_merged.groupby(df_merged.columns[4], dropna=False)[df_merged.columns[8]].transform("max") == df_merged[df_merged.columns[8]]]
    df_sorted_dates_violated = df_dates[df_dates[df_dates.columns[5]] != "عدم تخلف"].sort_values(df_dates.columns[8], ascending=True)
    df_sorted_dates_not_violated = df_dates[df_dates[df_dates.columns[5]] == "عدم تخلف"].sort_values(df_dates.columns[8], ascending=True)

    notary_list_not_visited = notary_list[notary_list_not_visited_index]
    empty_df = pd.DataFrame('', index=range(len(notary_list_not_visited)), columns=df_mistake_risk.columns)
    empty_df[empty_df.columns[4]] = notary_list_not_visited

    df_mistake_and_not_visited = pd.concat([df_mistake_risk, empty_df], axis=0)
    df_final = pd.concat([df_mistake_and_not_visited, df_sorted_dates_violated, df_sorted_dates_not_violated], axis=0)
    df_final.fillna("")

    df_org = df_final.reset_index(drop=True)
    df_org.insert(0, "شماره اصلی", value=list(range(1, df_org.shape[0] + 1)))
    df_org.rename(columns={df_org.columns[1]: "ردیف فایل مرجع"}, inplace=True)
    output_file_path = os.path.join(dir_name, "notaries_togo.xlsx")
    df_org.to_excel(output_file_path, index=False)

    workbook = openpyxl.load_workbook(output_file_path)
    sheet = workbook.active

    from openpyxl.styles import PatternFill, Font, Alignment

    sheet.sheet_view.rightToLeft = True

    max_column_index = sheet.max_column
    from openpyxl.utils import get_column_letter

    for column_index in range(1, max_column_index + 1):
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = 16

    sheet.column_dimensions["H"].width = 45

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    neon_green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    normal_blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    red_font = Font(color="FF0000")
    white_font = Font(color="FFFFFF")

    for col_index in [4, 10, 11]:
        for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index, max_row=sheet.max_row):
            for cell in row:
                if cell.value in ["Error", "Mistake Risk", "Mistake"]:
                    cell.fill = neon_green_fill
                    cell.font = red_font

    for row in sheet.iter_rows(min_row=len(df_mistake_risk) + 2, min_col=6, max_col=6, max_row=len(df_mistake_risk) + len(empty_df) + 1):
        for cell in row:
            cell.fill = normal_blue_fill
            cell.font = white_font

    for row in sheet.iter_rows(min_row=len(df_mistake_risk) + len(empty_df) + len(df_error) + 2, min_col=6, max_col=6, max_row=sheet.max_row):
        for cell in row:
            cell.fill = normal_blue_fill
            cell.font = white_font

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, readingOrder=2, vertical="center", horizontal="center")

    workbook.save(output_file_path)

    # Tkinter window to show success message
    root = tk.Tk()
    root.title("Success")
    root.geometry("400x200")
    tk.Label(root, text=success_label, fg="green", font=('Arial', 14)).pack(expand=True)
    tk.Label(root, text=success_label_1, fg="green", font=('Arial', 14)).pack(expand=True)
    root.after(3000, root.destroy)  # Auto-close after 3 seconds
    root.mainloop()

except Exception as e:
    # Tkinter window to show error message
    root = tk.Tk()
    root.title("Error")
    root.geometry("400x200")
    tk.Label(root, text=f"{str(e)} {error_label}", fg="red", font=('Arial', 14)).pack(expand=True)
    root.mainloop()
