# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import jdatetime
import openpyxl
import os 
from shutil import get_terminal_size
import time
columns = os.get_terminal_size().columns
import keyboard
from tkinter import ttk
from tkinter import *
import tkinter as tk
from tkinter import font
import sys
import warnings
import warnings

# Suppress all warnings (not recommended)
warnings.filterwarnings("ignore")
# Suppress the SettingWithCopyWarning
# pd.options.mode.chained_assignment = None 
pd.set_option("display.width",get_terminal_size()[0])
def welcome():
    s=[]
    s.append("--------------------نرم افزار تندار---------------")
    s.append("تعیین نوبت دفاتر اسناد رسمی جهت بازرسی")
    s.append("بهار 1403- نسخه 1.2")
    # s.append ("نسخه 1.1")
    s.append("-------------------------")
    s.append("مصطفی جلوه")
    s.append("Email: mostafaajelveh@gmail.com")
   
    
    root = tk.Tk()
    root.title("Farsi Text in Tkinter")

        # Set the geometry of the window
    root.geometry("800x600")
    root.config(bg="black")
        # Create a font that supports Farsi
    farsi_font = font.Font(family="Arial", size=14)
    columns = os.get_terminal_size().columns
    def label_initialization(farsi_texts):
    

        # Configure grid columns and rows for proper centering
        root.grid_columnconfigure(0, weight=1)  # Center column
        root.grid_columnconfigure(1, weight=1)  # Right column

            
        for num,i in enumerate(s):
            
            
            # padding = columns - len(i)
            # div_pd=padding//2
            if num<4:
                label = tk.Label(root, text=i, font=farsi_font,fg="#39FF14",bg="black")
                label.grid(row=num+4, column=0, columnspan=2, sticky='n', padx=10, pady=5)
                # print('\033[1m'+'033[38;5;46m'+'-' * (div_pd)+i+'-' * (div_pd))
            else:
                
                # print('\033[1m'+'\033[38;5;15m'+'-' * (div_pd)+i+'-' * (div_pd))
                label = tk.Label(root, text=i, font=farsi_font,fg="white",bg="black")
                label.grid(row=num+4, column=0, columnspan=2, sticky='n', padx=10, pady=5)
        for i in range(6,9):

            label = tk.Label(root, text="*"*50, font=farsi_font,fg="red",bg="black")
            label.grid(row=i+10, column=0, columnspan=2, sticky='n', padx=10, pady=5)
                
        # for i in range(4):   
        #     print()

        # for i in range(7):
        #     print('\033[1m'+"\033[38;5;196m"+"*"*columns)
        
        # time.sleep(5)
            


        
        
            
            # Create labels for each string and place them in the middle of the form
        # for i, text in enumerate(farsi_texts):
        #     label = tk.Label(root, text=text, font=farsi_font,fg="red",bg="black")
            
                
        #     label.grid(row=i, column=0, columnspan=2, sticky='n', padx=10, pady=5)


        # Run the application
    



    label_initialization(s)
    clear_button = tk.Button(root, text="ادامه", command=root.destroy, 
                            font=farsi_font, fg="black", bg="white", width=20)  # Increase button width to 20
    clear_button.grid(row=24, column=0, columnspan=2, sticky='n', padx=10, pady=40)

    root.mainloop()

def persian_to_georgian(persin_date):
    try:
        if persin_date!="" or None:
            date_list=persin_date.split("/")
            date_list=[int(i) for i in date_list]
            gregorian_date=jdatetime.date(date_list[0],date_list[1],date_list[2]).togregorian()
            return(gregorian_date)
    except Exception as e:
        return "Error"
def forward_fill_merged(df, merged_ranges):
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(':')
       

        
        
        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        end_row, end_col = openpyxl.utils.cell.coordinate_to_tuple(end_cell)

        
        start_row -= 2
        end_row -= 1
        start_col -=1  
        
        sub_df = df.iloc[start_row:end_row, start_col:end_col]
       
        
        sub_df = sub_df.ffill(axis=0)
        
        
        df.iloc[start_row:end_row, start_col:end_col] = sub_df
def search(event,combo_box,lst):
    global value
    value=event.widget.get()
    if value=="":
        combo_box["value"]=lst
    else:
        data=[]
        for item in lst:
            if value.lower() in item.lower():
                data.append(item)
        combo_box["value"]=data
def notary_adddress_concat(df1,df_address,base_df_col_num,sec_df_col_num):

    df_address=df_address.drop_duplicates(subset=df_address.columns[sec_df_col_num])
# Merge s1 with s based on matching values in s1["levels"] and s["class"]
# Include s index as a new column in s1 called "index_1"
    merged_df = df1.merge(df_address, left_on=[df1.columns[base_df_col_num]], right_on=[df_address.columns[sec_df_col_num]], how="left")
# merged_df = s1.merge(s.reset_index(), left_on="levels", right_on="class", how="left")
    # merged_df = merged_df.rename(columns={"index": "index_1"})
    return(merged_df)



def my_click():
    
    global combo_box1
    choosen1=combo_box1.get()
    root.destroy()
    file_path = dir_name+"\\"+"notaries_information_corrected_1403_07_23_example.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    
    
    sheet = workbook['Sheet1']  

   
    merged_cells_ranges = []

    
   
    for merged_range in sheet.merged_cells.ranges:
        merged_cells_ranges.append(merged_range.coord)
    workbook.close()
    column_dtype = {6: str,4:str}
    df = pd.read_excel(file_path, engine='openpyxl',usecols=list(range(0,10)),sheet_name="Sheet1",dtype=column_dtype)
    df[df.columns[6]]=df[df.columns[6]].fillna("")
    last_valid_index=df.iloc[:,6].last_valid_index()
    print("Last Row is:",last_valid_index+2)
    df=df.iloc[:last_valid_index+1,:]

    
    
    forward_fill_merged(df, merged_cells_ranges)
  

    grouped= df.groupby(df.columns[[0,1,2,3,4,5,7,8,9]].tolist(),dropna=False).apply(lambda x:x.index.min()+2,include_groups=False).reset_index()
    print(grouped,"sefef!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    
    grouped= df.groupby(df.columns[[0,1,2,3,4,5,7,8,9]].tolist(),dropna=False).apply(lambda x:x.index.min()+2).reset_index()
    print(grouped,"HAMED*****************")
    input()
    # df_merged = df.groupby(df.columns[[0,1,2,3]].tolist(),dropna=False)[df.columns[4]].apply(lambda x: ('*'.join(x))).reset_index()
    df_merged = df.groupby(df.columns[[0,1,2,3,4,5,7,8,9]].tolist(),dropna=False)[df.columns[6]].apply(lambda x: ('*'.join(x))).reset_index()
    df_merged["Excel_Index"]=grouped[grouped.columns[9]]
    df_merged=df_merged.iloc[:,[0,1,2,3,4,5,9,6,7,8,10]]
    # df_merged = df.groupby(['ردیف', 'نام', 'تاریخ'])['موارد'].apply(lambda x: ' '.join(x)).reset_index()
    pd.set_option("display.max_columns", None)
    # print(df_merged)
    # input("kamalll*************")
    df_merged_filtered=df_merged[df_merged[df_merged.columns[9]]==choosen1].copy()
    
    visited_notaries=np.array(df_merged_filtered[df_merged_filtered.columns[4]])
    
    # visited_notaries=np.unique(visited_notaries.astype(str))
   
    #Ta inja anjam dadam, be nazar ta inja file notary_information ra ba filtere shahr, shomarehaye daftarkhunasho dorost
    #mikhune be surate monhaser befard" emruuz 07 aban 1403

    # df_notary=pd.read_excel(dir_name+"\\"+"notary_list.xlsx",dtype={0:str})
   
    df_notary_filtered=df_notary[df_notary[df_notary.columns[8]]==choosen1]
    notary_list=np.array(df_notary_filtered[df_notary_filtered.columns[1]])
    # notary_list=np.unique(notary_list)
    np.set_printoptions(threshold=sys.maxsize)  
    

    notary_list_not_visited_index=np.invert(np.isin(notary_list,visited_notaries))

    notary_mistakes_index=np.invert(np.isin(visited_notaries,notary_list))



    # df_rule_cont=df_merged[df_merged[df_merged.columns[5]] !="عدم تخلف"]

    # print(df_rule_cont)
    
    df_merged_filtered[df_merged_filtered.columns[2]]=df_merged_filtered[df_merged_filtered.columns[2]].fillna("")

    df_merged_filtered.loc[:,"georgian"]=df_merged_filtered[df_merged_filtered.columns[2]].apply (lambda x:persian_to_georgian(x))
    


    df_error=df_merged_filtered[df_merged_filtered[df_merged_filtered.columns[11]]=="Error"]
    df_na=df_merged_filtered[df_merged_filtered[df_merged_filtered.columns[11]].isnull()].copy()
    df_na[df_na.columns[2]]="Mistake Risk"
    df_na_violated=df_na[df_na[df_na.columns[5]]!="عدم تخلف"]
    df_na_not_violated=df_na[df_na[df_na.columns[5]]=="عدم تخلف"]

   

    df_merged_filtered[df_merged_filtered.columns[11]]=pd.to_datetime(df_merged_filtered[df_merged_filtered.columns[11]],errors="coerce").dt.date
    
    df_mistake_risk=df_merged_filtered[notary_mistakes_index]
    df_mistake_risk["شماره دفترخانه اشتباه"]="Mistake"
  
    df_merged_filtered.dropna(subset=[df_merged_filtered.columns[11]], inplace=True)


    df_dates=df_merged_filtered[df_merged_filtered.groupby(df_merged_filtered.columns[4],dropna=False)[df_merged_filtered.columns[11]].transform("max")==df_merged_filtered[df_merged_filtered.columns[11]]]
    df_sorted_dates_violated=df_dates[df_dates[df_dates.columns[5]]!="عدم تخلف"].sort_values(df_dates.columns[11],ascending=True)
    df_sorted_dates_not_violated=df_dates[df_dates[df_dates.columns[5]]=="عدم تخلف"].sort_values(df_dates.columns[11],ascending=True)



    # df_merged.loc[notary_visited_mistakes_index,"شماره دفترخانه اشتباه"]="Mistake Risk"
    # pd.set_option('display.max_columns', None)

    notary_list_not_visited=notary_list[notary_list_not_visited_index]
    empty_df= pd.DataFrame('', index=range(len(notary_list_not_visited)), columns=df_mistake_risk.columns)
   
    empty_df[empty_df.columns[4]]=notary_list_not_visited
    empty_df[empty_df.columns[9]]=choosen1
    # print((empty_df).shape)
    # print((df_mistake_risk).shape)


    df_mistake_and_not_visited=pd.concat([df_mistake_risk,empty_df])
   
    df_mis_rows=df_mistake_risk.shape[0]
    empty_df_rows=empty_df.shape[0]
    df_error_rows=df_error.shape[0]
    df_na_violated_row=df_na_violated.shape[0]
    df_na_not_violated_row=df_na_not_violated.shape[0]
    df_sorted_dates_violated_row=df_sorted_dates_violated.shape[0]
    df_sorted_dates_not_violated_row=df_sorted_dates_not_violated.shape[0]
    df_org=pd.concat([df_mistake_and_not_visited,df_error,df_na_violated,df_sorted_dates_violated,df_na_not_violated,df_sorted_dates_not_violated])

 




    print(df_org.shape)

    df_org.insert(loc=0,column="ردیف اصلی",value=list(range(1,df_org.shape[0]+1)))
    df_org.rename(columns=({df_org.columns[1]:"ردیف فایل مرجع"}),inplace=True)
    df_address=pd.read_excel(dir_name+"\\"+"notary_address_corrected_FInal_1403_07_23_edited_08_12_final.xlsx",dtype={1:str})
    df_address_fl=df_address[df_address[df_address.columns[8]]==choosen1]
    df_org.fillna("")
    final_df=notary_adddress_concat(df_org,df_address_fl,5,1)
    print(df_org.shape)
    print(final_df.shape)
    input("SSAMMAMMAAMMA")

    output_file_path=dir_name+"\\"+"notaries_togo.xlsx"
    # Ta sare in ke mikham addresse daftarkhune ham ezafe konam. bebinam celule shomare daftarkhane khali ya
    # null bashe che kodi bayad bezanam!

    # for i in (range(df_org.shape[0])):
    #     if df_org.iloc[i,5].isnull()==False:#daghigan inja
    #         notary_number=df_org.iloc[i,5]
    #         notary_city=df_org.iloc[i,10]
    #         print(notary_number)
    #         print(notary_city)
    #         input()



    # df_org.to_excel(output_file_path,index=False)
    with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='w') as writer:
            final_df.to_excel(writer, index=False)
    # with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     # Write the new data to "Sheet2"
    #     new_data.to_excel(writer, sheet_name='Sheet2', index=False)
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill,Font
    workbook=openpyxl.load_workbook(output_file_path)
    sheet = workbook.active




    sheet.sheet_view.rightToLeft = True

    max_column_index = sheet.max_column

# Set the width of all active columns to 13
    for column_index in range(1, max_column_index + 1):
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = 16

    sheet.column_dimensions["H"].width = 45


    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    neon_green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    neon_blue_fill = PatternFill(start_color='1B03A3', end_color='1B03A3', fill_type='solid')
    normal_blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    neon_orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    neon_cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    neon_green_font = Font(color="00FF00")
    red_font = Font(color="FF0000")
    white_font = Font(color="FFFFFF")


    for col_index in [4,13,14]:
        for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index, max_row=sheet.max_row):
            for cell in row:
               
                if cell.value in ["Error", "Mistake Risk","Mistake"]:
                    # cell.fill = red_fill
                    cell.fill = neon_green_fill
                    # cell.font=neon_green_font
                    cell.font=red_font
                    first_col_cell = sheet.cell(row=cell.row, column=1)
                    first_col_cell.fill = red_fill
                    first_col_cell.font = white_font

    print(df_mis_rows+2)
    print(df_mis_rows+empty_df_rows+1)
    input("rows!!!!!!!!!!!!!!!") 
    for row in sheet.iter_rows(min_row=df_mis_rows+2, min_col=6, max_col=6, max_row=df_mis_rows+empty_df_rows+1):
        for cell in row:
                # cell.fill = red_fill
                cell.fill = normal_blue_fill
                # cell.font=neon_green_font
                cell.font=white_font
    print(df_error_rows,"llllllllllllllllllllllllllllllllllllllll")
    for row in sheet.iter_rows(min_row=df_mis_rows+empty_df_rows+df_error_rows+df_na_violated_row+2, min_col=6, max_col=6, max_row=df_mis_rows+empty_df_rows+df_error_rows+df_na_violated_row+df_sorted_dates_violated_row+1):
        for cell in row:
                # cell.fill = red_fill
                cell.fill = normal_blue_fill
                # cell.font=neon_green_font
                cell.font=white_font
    for row in sheet.iter_rows(min_row=df_mis_rows+empty_df_rows+df_error_rows+df_na_violated_row+df_sorted_dates_violated_row+df_na_not_violated_row+2, min_col=6, max_col=6, max_row=sheet.max_row):
        for cell in row:
                # cell.fill = red_fill
                cell.fill = normal_blue_fill
                # cell.font=neon_green_font
                cell.font=white_font

    for row in sheet.iter_rows():
        for cell in row:
        
            cell.alignment = Alignment(wrap_text=True,readingOrder=2,vertical="center",horizontal="center")

    workbook.save(output_file_path)
    workbook.close()
    print()
    print()
    print('\033[1m'+'\033[38;5;46m'+" "*(columns-len(success_label)-5),success_label)
    print('\033[1m'+'\033[38;5;46m'+" "*(columns-len(success_label_1)-5),success_label_1)

    print()
    print()
    print("hello")
try: 
    welcome()

    
    dir_name=os.path.dirname(os.path.abspath(__file__))
    os.system('cls' if os.name == 'nt' else 'clear')
    error_label=": شد مواجه خطا با فایل ایجاد"
    process_label="......[notary_visited_total.xlsx] فایل اطلاعات پردازش حال در"
    success_label=".شد ایجاد موفقیت با [notaries_togo.xlsx]  فایل"
    continue_label=".بفشارید را c دکمه انصراف برای یا و دهید فشار را Enter ادامه برای .شوید مطمئن جاری مسیر در [total.xlsx] فایل بودن موجود از ادامه از قبل "
    success_label_1=u'\u2713'+" "+"باشید موفق"
    
    df_notary=pd.read_excel(dir_name+"\\"+"notary_address_corrected_FInal_1403_07_23_edited_07_29.xlsx",dtype={1:str})
    df_new=df_notary.iloc[:,8].unique()
    df_new=list(df_new)
    root=Tk()
    root.title("بازرسی")
    root.geometry("800x600")
    label1=Label(root,text="شهر")
    label1.pack()
    combo_box1=ttk.Combobox(root,value=df_new)
    combo_box1.set("Search")
    combo_box1.pack()
    combo_box1.bind("<KeyRelease>",lambda event: search(event, combo_box1,df_new))

    my_button=Button(root,text="پردازش", command=lambda: my_click())
    my_button.pack()
    root.mainloop()
    
    
except Exception as e:
    error_label=": شد مواجه خطا با فایل ایجاد"
    print()
    print()
    print()
    print("\033[1m"+"\033[38;5;196m"+ " "*(columns-len(error_label)-len(str(e))-5),e,error_label)

finally:
    # Close the workbook in case of any exceptions
    if 'workbook' in locals():
        workbook.close()

