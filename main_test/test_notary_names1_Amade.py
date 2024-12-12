


# -*- coding: utf-8 -*-




import os
import pandas as pd
from openpyxl.styles import PatternFill,Font
from openpyxl.styles import Alignment
import openpyxl
from openpyxl.utils import get_column_letter
dir_name=os.path.dirname(os.path.abspath(__file__))
os.system('cls' if os.name == 'nt' else 'clear')
file_path = dir_name+"\\"+"notary_addresses.xlsx"
column_dtype = {0: str}
df = pd.read_excel(file_path, engine='openpyxl',usecols=list(range(0,7)),sheet_name="Sheet1",dtype=column_dtype)
df[df.columns[6]]=df[df.columns[6]].fillna("")
last_valid_index=df.iloc[:,6].last_valid_index()
print(df.tail())
print(last_valid_index)
notary_name=df[df.columns[6]]
print(notary_name)
number=[]
city=[]
state=[]
df["اشتباهات"]=''
for i,cell in enumerate(notary_name):
    splitted_cell=cell.split()
    print(splitted_cell)
    if "شماره" in splitted_cell:
        number.append(splitted_cell[splitted_cell.index("شماره")+1])
        if splitted_cell[splitted_cell.index("شماره")+1].isdigit()==False:
            df.iloc[i,7]="Mistake"
    else: 
        number.appned("Mistake")
    if "شهر" in splitted_cell:
        i=1
        city_var=""
        while splitted_cell[splitted_cell.index("شهر")+i]!="استان":
            if city_var=="":

                city_var=city_var+splitted_cell[splitted_cell.index("شهر")+i]
            else:
                city_var=city_var+" "+splitted_cell[splitted_cell.index("شهر")+i]
            i=i+1
        city.append(city_var)
    else:
        city.append("Mistake")
    if "استان" in splitted_cell:
        state_var=""
        for i in range(splitted_cell.index("استان")+1,len(splitted_cell)):
            if state_var=="":
                state_var=state_var+splitted_cell[i]
            else:
                state_var=state_var+" "+splitted_cell[i]
        
        state.append(state_var)

    else:
        state.append("Mistake")
print("gugul")
df["دفترخانه شماره"]=number
df["شهر"]=city
df["استان"]=state
pd.set_option('display.max_columns', None)
print(df.head())
print(df.tail())
print(df.shape)
output_file_path=dir_name+"\\"+"notary_address_corrected.xlsx"
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
neon_green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
neon_blue_fill = PatternFill(start_color='1B03A3', end_color='1B03A3', fill_type='solid')
normal_blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
neon_cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
neon_green_font = Font(color="00FF00")
red_font = Font(color="FF0000")
white_font = Font(color="FFFFFF")
alignment_center = Alignment(wrap_text=True, readingOrder=2, vertical="center", horizontal="center")

with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='w') as writer:
    df.to_excel(writer, index=False)

    workbook = writer.book
    sheet = workbook.active

    max_column_index = sheet.max_column

# Set the width of all active columns to 13
    for column_index in range(1, max_column_index + 1):
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = 16


    sheet.sheet_view.rightToLeft = True

    for col_index in [7, 8, 9, 10]:  # Adjust this range if needed based on the DataFrame's structure
            col_letter = get_column_letter(col_index + 1)  # Convert column index to Excel letter
            for row in range(2, sheet.max_row + 1):  # Iterate over rows, starting from the second row
                cell = sheet[f'{col_letter}{row}']
                if cell.value in ["Error", "Mistake Risk", "Mistake"]:
                    cell.fill = neon_green_fill
                    cell.font = red_font

        # Apply alignment to all cells
    for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = alignment_center
        